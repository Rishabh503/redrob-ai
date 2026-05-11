from openpyxl import load_workbook


def parse_jd_excel(filepath: str) -> list:
    """
    Reads Job Descriptions from an Excel file.

    Expected layout:
      Row 1-2  : Title / instructions (skipped)
      Row 3    : Headers (skipped)
      Row 4+   : One JD per row

    Columns:
      A — JD #
      B — Company
      C — Role
      D — Required Skills  (comma-separated canonical names)
      E — Preferred Skills (comma-separated canonical names)
    """
    wb = load_workbook(filepath, data_only=True)
    ws = wb.active

    job_descriptions = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        jd_id, company, role, required_raw, preferred_raw = row[:5]

        if not company or not role:
            continue  # skip empty rows

        required  = _split_skills(required_raw)
        preferred = _split_skills(preferred_raw)

        job_descriptions.append({
            "id":               str(jd_id).strip() if jd_id else "",
            "company":          str(company).strip(),
            "role":             str(role).strip(),
            "required_skills":  required,
            "preferred_skills": preferred,
        })

    print(f"  Loaded {len(job_descriptions)} JD(s) from '{filepath}'\n")
    return job_descriptions


def _split_skills(raw) -> list:
    """Splits a comma-separated skill string into a clean list."""
    if not raw:
        return []
    return [s.strip() for s in str(raw).split(",") if s.strip()]