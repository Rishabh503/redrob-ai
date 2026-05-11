import math
import sys
from openpyxl import load_workbook

# ── SKILL_ALIASES exactly as provided in the problem sheet ──────────────────
SKILL_ALIASES = {
    "python": "python", "pyhton": "python", "java": "java",
    "javascript": "javascript", "javascrpit": "javascript", "js": "javascript",
    "typescript": "typescript", "typescrpit": "typescript",
    "c++": "cpp", "cpp": "cpp", "r": "r", "kotlin": "kotlin",
    "machinelearning": "machine_learning", "machine learning": "machine_learning",
    "ml": "machine_learning", "sklearn": "machine_learning",
    "deeplearning": "deep_learning", "deep learning": "deep_learning",
    "deep-learning": "deep_learning",
    "tensorflow": "tensorflow", "pytorch": "pytorch", "keras": "keras",
    "nlp": "nlp", "bert": "bert", "xgboost": "xgboost",
    "feature engineering": "feature_engineering",
    "statistics": "statistics", "stats": "statistics",
    "regression": "regression", "clustering": "clustering",
    "data-viz": "data_visualization", "data visualization": "data_visualization",
    "data viz": "data_visualization", "matplotlib": "data_visualization",
    "tableau": "data_visualization", "power-bi": "data_visualization",
    "power bi": "data_visualization", "powerbi": "data_visualization",
    "pandas": "pandas", "numpy": "numpy",
    "react": "react", "reacts": "react", "reactjs": "react",
    "vue": "vue", "vue.js": "vue", "vuejs": "vue",
    "redux": "redux", "tailwind": "tailwind",
    "html/css": "html_css", "html css": "html_css", "html": "html_css", "css": "html_css",
    "jest": "jest", "graphql": "graphql",
    "node.js": "nodejs", "nodejs": "nodejs", "node js": "nodejs",
    "flask": "flask", "spring boot": "spring_boot", "springboot": "spring_boot",
    "rest api": "rest_api", "rest": "rest_api", "restapi": "rest_api",
    "microservices": "microservices",
    "sql": "sql", "mysql": "mysql", "mysq": "mysql",
    "postgresql": "postgresql", "postgres": "postgresql",
    "mongodb": "mongodb", "redis": "redis",
    "docker": "docker", "kubernetes": "kubernetes",
    "kubernates": "kubernetes", "k8s": "kubernetes",
    "ci/cd": "ci_cd", "cicd": "ci_cd", "ci cd": "ci_cd", "aws": "aws",
    "android": "android", "firebase": "firebase",
    "algorithms": "algorithms", "algoritms": "algorithms",
    "data structure": "data_structures", "data structures": "data_structures",
    "competitive programming": "competitive_programming",
    "ui/ux": "ui_ux", "ui ux": "ui_ux", "figma": "figma",
}

SORTED_ALIAS_KEYS = sorted(SKILL_ALIASES.keys(), key=len, reverse=True)


# ── Excel Parser ─────────────────────────────────────────────────────────────

def parse_jd_excel(filepath: str) -> list:
    """
    Reads JDs from an Excel file.
    Expected columns (row 3 = header, rows 4+ = data):
      Col A: JD #  | Col B: Company | Col C: Role
      Col D: Required Skills (comma-separated canonical names)
      Col E: Preferred Skills (comma-separated canonical names)
    """
    wb = load_workbook(filepath, data_only=True)
    ws = wb.active

    job_descriptions = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        jd_id, company, role, required_raw, preferred_raw = row[:5]

        # Skip empty rows
        if not company or not role:
            continue

        required  = [s.strip() for s in str(required_raw).split(",")  if s.strip()] if required_raw  else []
        preferred = [s.strip() for s in str(preferred_raw).split(",") if s.strip()] if preferred_raw else []

        job_descriptions.append({
            "id":               str(jd_id).strip() if jd_id else "",
            "company":          str(company).strip(),
            "role":             str(role).strip(),
            "required_skills":  required,
            "preferred_skills": preferred,
        })

    print(f"  Loaded {len(job_descriptions)} JD(s) from '{filepath}'\n")
    return job_descriptions


# ── Skill Normalization ───────────────────────────────────────────────────────

def normalize_skills(raw_skills: str) -> list:
    tokens = [t.strip().lower() for t in raw_skills.split(",")]
    canonical = []
    for token in tokens:
        for key in SORTED_ALIAS_KEYS:
            if token == key:
                canonical.append(SKILL_ALIASES[key])
                break
    seen, deduped = set(), []
    for skill in canonical:
        if skill not in seen:
            seen.add(skill)
            deduped.append(skill)
    return deduped


# ── Vocabulary ───────────────────────────────────────────────────────────────

def build_vocabulary(resume_datasets: list) -> list:
    vocab = set()
    for r in resume_datasets:
        vocab.update(normalize_skills(r["raw_skills"]))
    return sorted(vocab)


# ── TF-IDF ───────────────────────────────────────────────────────────────────

def compute_tfidf(resume: dict, vocabulary: list, resume_datasets: list) -> dict:
    skills = normalize_skills(resume["raw_skills"])
    n = len(skills)
    tfidf = {}
    for skill in vocabulary:
        if skill in skills:
            tf  = 1 / n
            df  = sum(1 for r in resume_datasets if skill in normalize_skills(r["raw_skills"]))
            idf = math.log(10 / df)
            tfidf[skill] = tf * idf
        else:
            tfidf[skill] = 0.0
    return tfidf


# ── JD Binary Vector ─────────────────────────────────────────────────────────

def build_jd_vector(job_description: dict, vocabulary: list) -> dict:
    present = set(job_description["required_skills"]) | set(job_description["preferred_skills"])
    return {skill: (1 if skill in present else 0) for skill in vocabulary}


# ── Cosine Similarity ────────────────────────────────────────────────────────

def cosine_similarity(tfidf: dict, jd_vector: dict, vocabulary: list) -> float:
    dot   = sum(tfidf[s] * jd_vector[s] for s in vocabulary)
    mag_a = math.sqrt(sum(tfidf[s] ** 2 for s in vocabulary))
    mag_b = math.sqrt(sum(jd_vector[s] ** 2 for s in vocabulary))
    return dot / (mag_a * mag_b) if mag_a and mag_b else 0.0


# ── Engine ───────────────────────────────────────────────────────────────────

class ResumeMatchingEngine:
    def __init__(self, resume_datasets, job_descriptions):
        self.resume_datasets  = resume_datasets
        self.job_descriptions = job_descriptions
        self.vocabulary       = build_vocabulary(resume_datasets)

    def rank(self, job_description: dict) -> list:
        jd_vec = build_jd_vector(job_description, self.vocabulary)
        scores = []
        for resume in self.resume_datasets:
            tfidf = compute_tfidf(resume, self.vocabulary, self.resume_datasets)
            score = cosine_similarity(tfidf, jd_vec, self.vocabulary)
            scores.append((resume["name"], score))
        scores.sort(key=lambda x: (-round(x[1], 2), x[0]))
        return scores[:3]

    def run(self):
        for jd in self.job_descriptions:
            label  = jd.get("id", "JD")
            top3   = self.rank(jd)
            result = ", ".join(f"{name}({score:.2f})" for name, score in top3)
            print(f"{label} — {jd['company']} ({jd['role']})")
            print(result)
            print()


# ── Resume Dataset (fixed) ───────────────────────────────────────────────────

resume_datasets = [
    {"name": "Arjun Sharma",    "raw_skills": "Pyhton, MachineLearning, SQL, pandas, numpy, Deep-learning"},
    {"name": "Priya Nair",      "raw_skills": "JavaScrpit, Reacts, Node.JS, MongoDb, REST api, HTML/CSS"},
    {"name": "Rahul Gupta",     "raw_skills": "Java, Spring Boot, MySql, Microservices, Docker, kubernates"},
    {"name": "Sneha Patel",     "raw_skills": "Python, TensorFlow, Keras, NLP, BERT, data-viz, matplotlib"},
    {"name": "Vikram Singh",    "raw_skills": "C++, Algoritms, Data Structure, competitive programming, python"},
    {"name": "Ananya Krishnan", "raw_skills": "javascript, vue.js, python, flask, PostgreSQL, AWS, CI/CD"},
    {"name": "Karan Mehta",     "raw_skills": "Python, Sklearn, XGboost, feature engineering, SQL, tableau"},
    {"name": "Deepika Rao",     "raw_skills": "Java, Android, Kotlin, Firebase, REST, UI/UX, figma"},
    {"name": "Aditya Kumar",    "raw_skills": "Reactjs, TypeScrpit, GraphQL, redux, tailwind, nodejs, jest"},
    {"name": "Meera Iyer",      "raw_skills": "python, R, statistics, ML, regression, clustering, Power-BI"},
]

# ── Entry Point ───────────────────────────────────────────────────────────────

if __name__ == "__main__":
    # Accept xlsx path as CLI arg, default to job_descriptions_input.xlsx
    xlsx_path = sys.argv[1] if len(sys.argv) > 1 else "job_descriptions_input.xlsx"

    print(f"Reading JDs from: {xlsx_path}")
    job_descriptions = parse_jd_excel(xlsx_path)

    if not job_descriptions:
        print("No JDs found in the Excel file. Please check the format.")
        sys.exit(1)

    engine = ResumeMatchingEngine(resume_datasets, job_descriptions)
    engine.run()